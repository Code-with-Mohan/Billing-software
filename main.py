import tkinter as tk
from tkinter import messagebox
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
import os
import datetime

class RestaurantBillingSoftware:
    def __init__(self, root):
        self.root = root
        self.root.title("Restaurant Billing Software")
        self.root.geometry("1000x700")
        self.root.configure(bg="#f2f2f2")

        # Load and set the background image
        self.background_image = Image.open("background.jpg")
        self.background_photo = ImageTk.PhotoImage(self.background_image)
        self.background_label = tk.Label(self.root, image=self.background_photo)
        self.background_label.place(relwidth=1, relheight=1)

        # Title
        self.title_label = tk.Label(self.root, text="Restaurant Billing Software", font=("Arial", 28, "bold"), bg="#34495e", fg="white", bd=12, relief=tk.GROOVE)
        self.title_label.grid(row=0, column=0, columnspan=4, sticky="ew")

        # Customer Info Frame
        self.customer_frame = tk.Frame(self.root, bd=10, relief=tk.GROOVE, bg="#ecf0f1", highlightbackground="black", highlightthickness=1)
        self.customer_frame.grid(row=1, column=0, columnspan=4, padx=20, pady=20, sticky="ew")

        self.customer_name_label = tk.Label(self.customer_frame, text="Customer Name:", font=("Arial", 14), bg="#ecf0f1")
        self.customer_name_label.grid(row=0, column=0, padx=20, pady=5, sticky="w")
        self.customer_name_entry = tk.Entry(self.customer_frame, font=("Arial", 14), width=25, bd=2, relief=tk.GROOVE)
        self.customer_name_entry.grid(row=0, column=1, padx=20, pady=5)

        self.customer_phone_label = tk.Label(self.customer_frame, text="Phone Number:", font=("Arial", 14), bg="#ecf0f1")
        self.customer_phone_label.grid(row=0, column=2, padx=20, pady=5, sticky="w")
        self.customer_phone_entry = tk.Entry(self.customer_frame, font=("Arial", 14), width=25, bd=2, relief=tk.GROOVE)
        self.customer_phone_entry.grid(row=0, column=3, padx=20, pady=5)

        # Menu Frame
        self.menu_frame = tk.Frame(self.root, bd=10, relief=tk.GROOVE, bg="#ecf0f1", highlightbackground="black", highlightthickness=1)
        self.menu_frame.grid(row=2, column=0, rowspan=4, padx=20, pady=20, sticky="nsew")

        self.menu_title = tk.Label(self.menu_frame, text="Menu", font=("Arial", 18, "bold"), bg="#ecf0f1")
        self.menu_title.grid(row=0, columnspan=3, pady=10)

        # Sample menu items (you can add more items as needed)
        self.menu_items = [
            {"name": "Burger", "price": 120.00},
            {"name": "Pizza", "price": 250.00},
            {"name": "Pasta", "price": 200.00},
            {"name": "Salad", "price": 100.00},
            {"name": "Soda", "price": 50.00},
        ]

        self.menu_vars = []
        self.quantity_vars = []
        for index, item in enumerate(self.menu_items):
            var = tk.StringVar(value=item["name"])
            self.menu_vars.append(var)
            quantity_var = tk.IntVar(value=0)
            self.quantity_vars.append(quantity_var)

            name_label = tk.Label(self.menu_frame, text=item["name"], font=("Arial", 14), bg="#ecf0f1")
            name_label.grid(row=index + 1, column=0, padx=20, pady=10, sticky="w")

            price_label = tk.Label(self.menu_frame, text=f"₹{item['price']:.2f}", font=("Arial", 14), bg="#ecf0f1")
            price_label.grid(row=index + 1, column=1, padx=20, pady=10)

            quantity_spinbox = tk.Spinbox(self.menu_frame, from_=0, to=100, textvariable=quantity_var, font=("Arial", 14), width=5, bd=2, relief=tk.GROOVE)
            quantity_spinbox.grid(row=index + 1, column=2, padx=20, pady=10)

        # Buttons Frame
        self.buttons_frame = tk.Frame(self.root, bg="#f2f2f2")
        self.buttons_frame.grid(row=2, column=1, padx=20, pady=20, sticky="nsew")

        button_style = {"font": ("Arial", 14), "fg": "white", "bd": 4, "relief": tk.RAISED}

        self.add_item_button = tk.Button(self.buttons_frame, text="Add Item", command=self.add_item, **button_style, bg="#34495e", width=15)
        self.add_item_button.grid(row=0, column=0, pady=10)

        self.generate_bill_button = tk.Button(self.buttons_frame, text="Generate Bill", command=self.generate_bill, **button_style, bg="#34495e", width=15)
        self.generate_bill_button.grid(row=1, column=0, pady=10)

        self.clear_button = tk.Button(self.buttons_frame, text="Clear", command=self.clear_bill, **button_style, bg="#e74c3c", width=15)
        self.clear_button.grid(row=2, column=0, pady=10)

        self.save_excel_button = tk.Button(self.buttons_frame, text="Save as Excel", command=self.save_as_excel, **button_style, bg="#2980b9", width=15)
        self.save_excel_button.grid(row=3, column=0, pady=10)

        # Bill Frame
        self.bill_frame = tk.Frame(self.root, bd=10, relief=tk.GROOVE, bg="#ecf0f1", highlightbackground="black", highlightthickness=1)
        self.bill_frame.grid(row=2, column=2, columnspan=2, rowspan=4, padx=20, pady=20, sticky="nsew")

        self.bill_text = tk.Text(self.bill_frame, font=("Arial", 12), bg="#ffffff", bd=2, relief=tk.GROOVE)
        self.bill_text.pack(fill=tk.BOTH, expand=True)

        # Variables for bill calculation
        self.items = []
        self.total_price = 0.0

        # Grid configuration for responsiveness
        self.root.grid_columnconfigure(0, weight=1)
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_columnconfigure(2, weight=1)
        self.root.grid_columnconfigure(3, weight=1)
        self.root.grid_rowconfigure(2, weight=1)
        self.root.grid_rowconfigure(3, weight=1)
        self.root.grid_rowconfigure(4, weight=1)
        self.root.grid_rowconfigure(5, weight=1)

    def add_item(self):
        self.bill_text.delete(1.0, tk.END)
        self.total_price = 0.0
        self.items.clear()
        for index, item in enumerate(self.menu_items):
            quantity = self.quantity_vars[index].get()
            if quantity > 0:
                total_item_price = quantity * item["price"]
                self.items.append((item["name"], quantity, item["price"], total_item_price))
                self.bill_text.insert(tk.END, f"{item['name']}\t{quantity}\t₹{item['price']:.2f}\t₹{total_item_price:.2f}\n")
                self.total_price += total_item_price

    def generate_bill(self):
        customer_name = self.customer_name_entry.get()
        customer_phone = self.customer_phone_entry.get()
        self.bill_text.insert(tk.END, "\n----------------------------------------\n")
        self.bill_text.insert(tk.END, f"Customer Name: {customer_name}\n")
        self.bill_text.insert(tk.END, f"Phone Number: {customer_phone}\n")
        self.bill_text.insert(tk.END, "----------------------------------------\n")
        self.add_item()
        self.bill_text.insert(tk.END, "----------------------------------------\n")
        self.bill_text.insert(tk.END, f"Total Price: \t\t\t₹{self.total_price:.2f}\n")
        self.bill_text.insert(tk.END, "----------------------------------------\n")

    def clear_bill(self):
        self.bill_text.delete(1.0, tk.END)
        self.clear_quantities()

    def clear_quantities(self):
        for quantity_var in self.quantity_vars:
            quantity_var.set(0)

    def save_as_excel(self):
        if not self.items:
            messagebox.showerror("Error", "No items to save.")
            return

        customer_name = self.customer_name_entry.get()
        customer_phone = self.customer_phone_entry.get()

        filename = "bills.xlsx"
        file_exists = os.path.exists(filename)
        if file_exists:
            workbook = load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Bills"
            sheet.append(["Order Number", "Customer Name", "Phone Number", "Item", "Quantity", "Price", "Total"])

        order_number = sheet.max_row  # Use max_row as the order number (incrementing with each order)

        for item in self.items:
            sheet.append([order_number, customer_name, customer_phone, item[0], item[1], item[2], item[3]])
        
        sheet.append([order_number, customer_name, customer_phone, "Total Price", "", "", self.total_price])

        workbook.save(filename)
        messagebox.showinfo("Success", f"Bill saved as Excel: {filename}")

if __name__ == "__main__":
    root = tk.Tk()
    app = RestaurantBillingSoftware(root)
    root.mainloop()
